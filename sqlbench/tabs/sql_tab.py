"""SQL utility tab."""

import re
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import threading


class SQLTab:
    def __init__(self, parent, app, connection, conn_name, version, adapter):
        self.app = app
        self.connection = connection
        self.conn_name = conn_name
        self.version = version
        self.adapter = adapter
        self.frame = ttk.Frame(parent)
        self.frame.pack(fill=tk.BOTH, expand=True)
        self._running = False
        self._cursor = None

        # Pagination state
        self._all_rows = []
        self._columns = []
        self._column_info = []  # Full column metadata
        self._current_page = 0
        self._rows_per_page = 1000
        self._show_all = False
        self._total_rows = 0  # Total rows in result set (from COUNT)
        self._base_sql = ""   # Original SQL without pagination
        self._sort_column = None  # Current sort column
        self._sort_reverse = False  # Sort direction
        self._results_base_widths = {}  # Base column widths (at font size 10)

        self._create_widgets()
        self._bind_keys()

    def _bind_keys(self):
        # Ensure standard text editing shortcuts work
        self.sql_text.bind("<Control-a>", self._select_all)
        self.sql_text.bind("<Control-A>", self._select_all)
        self.sql_text.bind("<Control-Shift-f>", lambda e: self._format_sql())
        self.sql_text.bind("<Control-Shift-F>", lambda e: self._format_sql())

    def _select_all(self, event=None):
        """Select all text in SQL editor."""
        self.sql_text.tag_add("sel", "1.0", "end-1c")
        self.sql_text.mark_set("insert", "end-1c")
        return "break"

    def _setup_syntax_highlighting(self):
        """Configure syntax highlighting tags for the SQL editor."""
        is_dark = self.app.dark_mode_var.get()

        if is_dark:
            keyword_color = "#CC7832"      # Orange for keywords
            function_color = "#FFC66D"     # Yellow for functions
            string_color = "#6A8759"       # Green for strings
            comment_color = "#808080"      # Gray for comments
            number_color = "#6897BB"       # Blue for numbers
            operator_color = "#A9B7C6"     # Light gray for operators
        else:
            keyword_color = "#0000FF"      # Blue for keywords
            function_color = "#795E26"     # Brown for functions
            string_color = "#008000"       # Green for strings
            comment_color = "#808080"      # Gray for comments
            number_color = "#098658"       # Teal for numbers
            operator_color = "#000000"     # Black for operators

        self.sql_text.tag_configure("keyword", foreground=keyword_color, font=("TkFixedFont", self.app.font_size, "bold"))
        self.sql_text.tag_configure("function", foreground=function_color)
        self.sql_text.tag_configure("string", foreground=string_color)
        self.sql_text.tag_configure("comment", foreground=comment_color, font=("TkFixedFont", self.app.font_size, "italic"))
        self.sql_text.tag_configure("number", foreground=number_color)
        self.sql_text.tag_configure("operator", foreground=operator_color)

        # SQL keywords
        self._sql_keywords = {
            'SELECT', 'FROM', 'WHERE', 'AND', 'OR', 'NOT', 'IN', 'IS', 'NULL',
            'JOIN', 'LEFT', 'RIGHT', 'INNER', 'OUTER', 'FULL', 'CROSS', 'ON',
            'GROUP', 'BY', 'HAVING', 'ORDER', 'ASC', 'DESC', 'LIMIT', 'OFFSET',
            'INSERT', 'INTO', 'VALUES', 'UPDATE', 'SET', 'DELETE', 'CREATE',
            'TABLE', 'INDEX', 'VIEW', 'DROP', 'ALTER', 'ADD', 'COLUMN',
            'PRIMARY', 'KEY', 'FOREIGN', 'REFERENCES', 'CONSTRAINT', 'UNIQUE',
            'DEFAULT', 'CHECK', 'CASCADE', 'RESTRICT', 'UNION', 'ALL', 'DISTINCT',
            'AS', 'CASE', 'WHEN', 'THEN', 'ELSE', 'END', 'LIKE', 'BETWEEN',
            'EXISTS', 'ANY', 'SOME', 'FETCH', 'FIRST', 'NEXT', 'ROWS', 'ONLY',
            'WITH', 'RECURSIVE', 'OVER', 'PARTITION', 'ROW_NUMBER', 'RANK',
            'TRUE', 'FALSE', 'BEGIN', 'COMMIT', 'ROLLBACK', 'TRANSACTION',
            'TRUNCATE', 'GRANT', 'REVOKE', 'CALL', 'DECLARE', 'CURSOR', 'FOR'
        }

        # SQL functions
        self._sql_functions = {
            'COUNT', 'SUM', 'AVG', 'MIN', 'MAX', 'ABS', 'ROUND', 'FLOOR', 'CEIL',
            'CEILING', 'COALESCE', 'NULLIF', 'IFNULL', 'NVL', 'CAST', 'CONVERT',
            'UPPER', 'LOWER', 'TRIM', 'LTRIM', 'RTRIM', 'LENGTH', 'LEN', 'SUBSTR',
            'SUBSTRING', 'CONCAT', 'REPLACE', 'INSTR', 'LOCATE', 'POSITION',
            'LEFT', 'RIGHT', 'LPAD', 'RPAD', 'REVERSE', 'REPEAT',
            'DATE', 'TIME', 'TIMESTAMP', 'YEAR', 'MONTH', 'DAY', 'HOUR', 'MINUTE',
            'SECOND', 'NOW', 'CURRENT_DATE', 'CURRENT_TIME', 'CURRENT_TIMESTAMP',
            'DATEADD', 'DATEDIFF', 'EXTRACT', 'TO_DATE', 'TO_CHAR', 'TO_NUMBER',
            'DENSE_RANK', 'NTILE', 'LAG', 'LEAD', 'FIRST_VALUE', 'LAST_VALUE',
            'LISTAGG', 'STRING_AGG', 'GROUP_CONCAT', 'JSON_VALUE', 'JSON_QUERY'
        }

    def _on_sql_key_release(self, event=None):
        """Handle key release in SQL editor - debounced highlighting."""
        # Cancel any pending highlight job
        if self._highlight_job:
            self.sql_text.after_cancel(self._highlight_job)
        # Schedule highlighting after 150ms of no typing
        self._highlight_job = self.sql_text.after(150, self._highlight_sql)

    def _highlight_sql(self):
        """Apply syntax highlighting to the SQL text."""
        self._highlight_job = None

        # Remove all existing tags
        for tag in ("keyword", "function", "string", "comment", "number", "operator"):
            self.sql_text.tag_remove(tag, "1.0", "end")

        text = self.sql_text.get("1.0", "end-1c")
        if not text.strip():
            return

        # Order matters: strings and comments first (so keywords inside them aren't highlighted)

        # Highlight single-line comments (-- ...)
        for match in re.finditer(r'--[^\n]*', text):
            self._apply_tag("comment", match.start(), match.end())

        # Highlight multi-line comments (/* ... */)
        for match in re.finditer(r'/\*[\s\S]*?\*/', text):
            self._apply_tag("comment", match.start(), match.end())

        # Highlight strings (single quotes)
        for match in re.finditer(r"'(?:[^'\\]|\\.)*'", text):
            self._apply_tag("string", match.start(), match.end())

        # Highlight numbers
        for match in re.finditer(r'\b\d+\.?\d*\b', text):
            if not self._is_inside_string_or_comment(match.start()):
                self._apply_tag("number", match.start(), match.end())

        # Highlight keywords and functions (word boundaries)
        for match in re.finditer(r'\b[A-Za-z_][A-Za-z0-9_]*\b', text):
            word = match.group().upper()
            start = match.start()
            if self._is_inside_string_or_comment(start):
                continue
            if word in self._sql_keywords:
                self._apply_tag("keyword", start, match.end())
            elif word in self._sql_functions:
                self._apply_tag("function", start, match.end())

    def _apply_tag(self, tag, start_char, end_char):
        """Apply a tag to a character range."""
        start_idx = f"1.0+{start_char}c"
        end_idx = f"1.0+{end_char}c"
        self.sql_text.tag_add(tag, start_idx, end_idx)

    def _is_inside_string_or_comment(self, char_pos):
        """Check if a character position is inside a string or comment."""
        idx = f"1.0+{char_pos}c"
        tags = self.sql_text.tag_names(idx)
        return "string" in tags or "comment" in tags

    def _format_sql(self):
        """Format the SQL statement with consistent indentation and line breaks."""
        sql = self.sql_text.get("1.0", "end-1c").strip()
        if not sql:
            return

        # Preserve cursor position roughly
        try:
            cursor_pos = self.sql_text.index("insert")
        except Exception:
            cursor_pos = "1.0"

        formatted = self._do_format_sql(sql)

        # Update text
        self.sql_text.delete("1.0", "end")
        self.sql_text.insert("1.0", formatted)

        # Restore cursor (roughly)
        try:
            self.sql_text.mark_set("insert", cursor_pos)
            self.sql_text.see("insert")
        except Exception:
            pass

        # Re-highlight
        self._highlight_sql()
        self.app.statusbar.config(text="SQL formatted")

    def _do_format_sql(self, sql):
        """Perform SQL formatting."""
        # Remove extra whitespace
        sql = re.sub(r'\s+', ' ', sql).strip()

        # Keywords that should start a new line
        newline_before = [
            'SELECT', 'FROM', 'WHERE', 'AND', 'OR', 'ORDER BY', 'GROUP BY',
            'HAVING', 'LIMIT', 'OFFSET', 'JOIN', 'LEFT JOIN', 'RIGHT JOIN',
            'INNER JOIN', 'OUTER JOIN', 'FULL JOIN', 'CROSS JOIN', 'ON',
            'UNION', 'UNION ALL', 'INSERT INTO', 'VALUES', 'UPDATE', 'SET',
            'DELETE FROM', 'CREATE TABLE', 'CREATE INDEX', 'CREATE VIEW',
            'DROP TABLE', 'ALTER TABLE', 'FETCH FIRST', 'WITH'
        ]

        # Sort by length (longest first) to match multi-word keywords first
        newline_before.sort(key=len, reverse=True)

        result = sql

        # Add newlines before major keywords
        for keyword in newline_before:
            # Pattern to match keyword (case insensitive, word boundary)
            pattern = r'(?i)(?<!\n)\s+\b(' + re.escape(keyword) + r')\b'
            result = re.sub(pattern, r'\n\1', result)

        # Indent lines after SELECT, with commas on separate lines
        lines = result.split('\n')
        formatted_lines = []
        indent = 0

        for line in lines:
            line = line.strip()
            if not line:
                continue

            line_upper = line.upper()

            # Adjust indentation
            if line_upper.startswith(('SELECT', 'FROM', 'WHERE', 'GROUP BY', 'ORDER BY', 'HAVING')):
                indent = 0
            elif line_upper.startswith(('AND', 'OR')):
                indent = 1
            elif line_upper.startswith(('JOIN', 'LEFT', 'RIGHT', 'INNER', 'OUTER', 'FULL', 'CROSS', 'ON')):
                indent = 1
            elif line_upper.startswith(('UNION',)):
                indent = 0

            formatted_lines.append('    ' * indent + line)

        return '\n'.join(formatted_lines)

    def _add_tooltip(self, widget, text):
        """Add a tooltip to a widget."""
        tooltip = None

        def show(event):
            nonlocal tooltip
            x, y, _, _ = widget.bbox("insert") if hasattr(widget, 'bbox') else (0, 0, 0, 0)
            x += widget.winfo_rootx() + 25
            y += widget.winfo_rooty() + 25
            tooltip = tk.Toplevel(widget)
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{x}+{y}")
            label = tk.Label(tooltip, text=text, background="#ffffe0", relief="solid", borderwidth=1, padx=4, pady=2)
            label.pack()

        def hide(event):
            nonlocal tooltip
            if tooltip:
                tooltip.destroy()
                tooltip = None

        widget.bind("<Enter>", show)
        widget.bind("<Leave>", hide)

    def _create_widgets(self):
        # Top button bar
        btn_frame = ttk.Frame(self.frame)
        btn_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

        self.run_btn = ttk.Button(btn_frame, text="Run", command=self._run_query)
        self.run_btn.pack(side=tk.LEFT, padx=2)
        self._add_tooltip(self.run_btn, "Run query (F5)")

        self.cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self._cancel_query, state=tk.DISABLED)
        self.cancel_btn.pack(side=tk.LEFT, padx=2)
        self._add_tooltip(self.cancel_btn, "Cancel query (Esc)")

        save_btn = ttk.Button(btn_frame, text="Save Query", command=self._save_query)
        save_btn.pack(side=tk.LEFT, padx=2)
        self._add_tooltip(save_btn, "Save query to file (Ctrl+S)")

        load_btn = ttk.Button(btn_frame, text="Load Query", command=self._load_query)
        load_btn.pack(side=tk.LEFT, padx=2)
        self._add_tooltip(load_btn, "Load query from file (Ctrl+O)")

        ttk.Button(btn_frame, text="Clear", command=self._clear).pack(side=tk.LEFT, padx=2)

        ttk.Separator(btn_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)

        format_btn = ttk.Button(btn_frame, text="Format", command=self._format_sql)
        format_btn.pack(side=tk.LEFT, padx=2)
        self._add_tooltip(format_btn, "Format SQL (Ctrl+Shift+F)")

        # Connection info label
        ttk.Label(btn_frame, text=f"  [{self.conn_name}]").pack(side=tk.RIGHT, padx=5)

        # Paned window for SQL entry and results
        self.paned = ttk.PanedWindow(self.frame, orient=tk.VERTICAL)
        self.paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # SQL entry area
        sql_frame = ttk.LabelFrame(self.paned, text="SQL Statement")
        self.paned.add(sql_frame, weight=1)

        font_spec = ("TkFixedFont", self.app.font_size)
        self.sql_text = tk.Text(sql_frame, height=10, wrap=tk.NONE, font=font_spec)
        sql_scroll_y = ttk.Scrollbar(sql_frame, orient=tk.VERTICAL, command=self.sql_text.yview)
        sql_scroll_x = ttk.Scrollbar(sql_frame, orient=tk.HORIZONTAL, command=self.sql_text.xview)
        self.sql_text.configure(yscrollcommand=sql_scroll_y.set, xscrollcommand=sql_scroll_x.set)

        sql_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        sql_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.sql_text.pack(fill=tk.BOTH, expand=True)

        # Setup syntax highlighting
        self._setup_syntax_highlighting()
        self._highlight_job = None  # For debounced highlighting
        self.sql_text.bind("<KeyRelease>", self._on_sql_key_release)

        # Create context menu for SQL editor
        self._create_sql_context_menu()

        # SQL editor search state
        self._sql_search_matches = []
        self._sql_search_index = -1
        self._sql_last_search = ""
        self.sql_text.tag_configure("search_highlight", background="#FFFF00", foreground="#000000")
        self.sql_text.tag_configure("search_current", background="#FF8C00", foreground="#000000")

        # Results area with tabs
        results_frame = ttk.Frame(self.paned)
        self.paned.add(results_frame, weight=2)

        # Create notebook for Results/Fields/Statistics tabs
        self.results_notebook = ttk.Notebook(results_frame)
        self.results_notebook.pack(fill=tk.BOTH, expand=True)

        # Tab 1: Results
        self._create_results_tab()

        # Tab 2: Fields
        self._create_fields_tab()

        # Tab 3: Statistics
        self._create_statistics_tab()

    def _create_results_tab(self):
        """Create the Results tab with data grid and pagination."""
        results_tab = ttk.Frame(self.results_notebook)
        self.results_notebook.add(results_tab, text="Results")

        # Pagination controls at top
        self._create_pagination_controls(results_tab)

        # Status label at bottom (pack before tree so it stays at bottom)
        self.page_label = ttk.Label(results_tab, text="No results")
        self.page_label.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=3)

        # Results treeview
        tree_frame = ttk.Frame(results_tab)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        self.results_tree = ttk.Treeview(tree_frame, show="headings")
        results_scroll_y = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.results_tree.yview)
        results_scroll_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.results_tree.xview)
        self.results_tree.configure(yscrollcommand=results_scroll_y.set, xscrollcommand=results_scroll_x.set)

        results_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        results_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.results_tree.pack(fill=tk.BOTH, expand=True)

        # Double-click handling (header separator vs data row)
        self.results_tree.bind("<Double-1>", self._on_results_double_click)

    def _create_fields_tab(self):
        """Create the Fields tab showing column metadata."""
        fields_tab = ttk.Frame(self.results_notebook)
        self.results_notebook.add(fields_tab, text="Fields")

        # Fields treeview
        columns = ("table", "name", "type", "display_size", "precision", "scale", "nullable")
        self.fields_tree = ttk.Treeview(fields_tab, columns=columns, show="headings")

        self.fields_tree.heading("table", text="Table")
        self.fields_tree.heading("name", text="Column Name")
        self.fields_tree.heading("type", text="Type")
        self.fields_tree.heading("display_size", text="Display Size")
        self.fields_tree.heading("precision", text="Precision")
        self.fields_tree.heading("scale", text="Scale")
        self.fields_tree.heading("nullable", text="Nullable")

        # Base column widths (at font size 10)
        self._fields_base_widths = {
            "table": 120, "name": 150, "type": 120,
            "display_size": 100, "precision": 80, "scale": 80, "nullable": 80
        }
        scale = self.app.font_size / 10.0
        for col, base_w in self._fields_base_widths.items():
            self.fields_tree.column(col, width=int(base_w * scale))

        fields_scroll_y = ttk.Scrollbar(fields_tab, orient=tk.VERTICAL, command=self.fields_tree.yview)
        self.fields_tree.configure(yscrollcommand=fields_scroll_y.set)

        fields_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.fields_tree.pack(fill=tk.BOTH, expand=True)

    def _create_statistics_tab(self):
        """Create the Statistics tab showing query execution info."""
        stats_tab = ttk.Frame(self.results_notebook)
        self.results_notebook.add(stats_tab, text="Statistics")

        # Statistics text area
        self.stats_text = tk.Text(stats_tab, wrap=tk.WORD, font=("Courier", self.app.font_size))
        stats_scroll_y = ttk.Scrollbar(stats_tab, orient=tk.VERTICAL, command=self.stats_text.yview)
        self.stats_text.configure(yscrollcommand=stats_scroll_y.set)

        stats_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.stats_text.pack(fill=tk.BOTH, expand=True)

        self.stats_text.insert("1.0", "Execute a query to see statistics.")
        self.stats_text.config(state=tk.DISABLED)

    def _create_pagination_controls(self, parent):
        """Create pagination controls above results."""
        self.paging_frame = ttk.Frame(parent)
        self.paging_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=3)

        # Left side - navigation buttons
        nav_frame = ttk.Frame(self.paging_frame)
        nav_frame.pack(side=tk.LEFT)

        self.first_btn = ttk.Button(nav_frame, text="◀◀", width=3, command=self._first_page, state=tk.DISABLED)
        self.first_btn.pack(side=tk.LEFT, padx=1)
        self.prev_btn = ttk.Button(nav_frame, text="◀", width=3, command=self._prev_page, state=tk.DISABLED)
        self.prev_btn.pack(side=tk.LEFT, padx=1)
        self.next_btn = ttk.Button(nav_frame, text="▶", width=3, command=self._next_page, state=tk.DISABLED)
        self.next_btn.pack(side=tk.LEFT, padx=1)
        self.last_btn = ttk.Button(nav_frame, text="▶▶", width=3, command=self._last_page, state=tk.DISABLED)
        self.last_btn.pack(side=tk.LEFT, padx=1)

        # Save To dropdown
        self.save_menu = tk.Menu(nav_frame, tearoff=0)
        self.save_menu.add_command(label="Copy to Clipboard", command=self._copy_to_clipboard)
        self.save_menu.add_separator()
        self.save_menu.add_command(label="Excel (.xlsx)", command=self._save_to_excel)
        self.save_menu.add_command(label="CSV (.csv)", command=self._save_to_csv)
        self.save_menu.add_command(label="JSON (.json)", command=self._save_to_json)
        self.save_btn = ttk.Menubutton(nav_frame, text="Save To", menu=self.save_menu)
        self.save_btn.pack(side=tk.LEFT, padx=(15, 0))

        # Results search frame
        results_search_frame = ttk.Frame(nav_frame)
        results_search_frame.pack(side=tk.LEFT, padx=(20, 0))

        ttk.Label(results_search_frame, text="Search:").pack(side=tk.LEFT, padx=(0, 2))
        self.results_search_var = tk.StringVar()
        self.results_search_entry = ttk.Entry(results_search_frame, textvariable=self.results_search_var, width=15)
        self.results_search_entry.pack(side=tk.LEFT, padx=2)
        self.results_search_entry.bind("<Return>", lambda e: self._results_search_next())
        self.results_search_entry.bind("<Shift-Return>", lambda e: self._results_search_prev())

        self.results_search_prev_btn = ttk.Button(results_search_frame, text="<", width=2, command=self._results_search_prev)
        self.results_search_prev_btn.pack(side=tk.LEFT, padx=1)
        self.results_search_next_btn = ttk.Button(results_search_frame, text=">", width=2, command=self._results_search_next)
        self.results_search_next_btn.pack(side=tk.LEFT, padx=1)

        self._results_search_matches = []
        self._results_search_index = -1
        self._results_last_search = ""

        # Right side - rows per page and show all
        settings_frame = ttk.Frame(self.paging_frame)
        settings_frame.pack(side=tk.RIGHT)

        # Show all checkbox
        self.show_all_var = tk.BooleanVar(value=False)
        self.show_all_cb = ttk.Checkbutton(
            settings_frame, text="Show All",
            variable=self.show_all_var,
            command=self._on_show_all_changed
        )
        self.show_all_cb.pack(side=tk.RIGHT, padx=10)

        # Rows per page
        ttk.Label(settings_frame, text="Rows per page:").pack(side=tk.LEFT, padx=(0, 5))
        self.rows_per_page_var = tk.StringVar(value="1000")
        self.rows_per_page_spin = ttk.Spinbox(
            settings_frame, from_=100, to=10000, increment=100,
            width=7, textvariable=self.rows_per_page_var
        )
        self.rows_per_page_spin.pack(side=tk.LEFT)
        self.rows_per_page_spin.bind("<Return>", self._on_rows_per_page_changed)
        self.rows_per_page_spin.bind("<FocusOut>", self._on_rows_per_page_changed)
        self.rows_per_page_spin.bind("<<Increment>>", self._on_rows_per_page_changed)
        self.rows_per_page_spin.bind("<<Decrement>>", self._on_rows_per_page_changed)

    def _on_show_all_changed(self):
        """Handle show all checkbox change - re-run query with new limit."""
        self._show_all = self.show_all_var.get()
        self._current_page = 0

        # If we have a query and results, re-run to fetch with new limit
        sql = self.sql_text.get("1.0", tk.END).strip()
        if sql and self._all_rows:
            self._run_query()
        else:
            self._refresh_display()

    def _on_rows_per_page_changed(self, event=None):
        """Handle rows per page change."""
        try:
            new_value = int(self.rows_per_page_var.get())
            if new_value < 1:
                new_value = 100
            self._rows_per_page = new_value
            self._current_page = 0
            if self._base_sql:
                self._fetch_page()
        except ValueError:
            self.rows_per_page_var.set(str(self._rows_per_page))

    def _first_page(self):
        """Go to first page."""
        if self._current_page != 0:
            self._current_page = 0
            self._fetch_page()

    def _prev_page(self):
        """Go to previous page."""
        if self._current_page > 0:
            self._current_page -= 1
            self._fetch_page()

    def _next_page(self):
        """Go to next page."""
        max_page = self._get_max_page()
        if self._current_page < max_page:
            self._current_page += 1
            self._fetch_page()

    def _last_page(self):
        """Go to last page."""
        max_page = self._get_max_page()
        if self._current_page != max_page:
            self._current_page = max_page
            self._fetch_page()

    def _get_max_page(self):
        """Get maximum page index."""
        if self._total_rows <= 0 or self._rows_per_page <= 0:
            return 0
        return max(0, (self._total_rows - 1) // self._rows_per_page)

    def _fetch_page(self):
        """Fetch the current page of results from the server."""
        if not self._base_sql or self._running:
            return

        self._set_running(True)
        self.app.statusbar.config(text=f"Fetching page {self._current_page + 1}...")

        thread = threading.Thread(target=self._execute_page_query, daemon=True)
        thread.start()

    def _execute_page_query(self):
        """Execute query for current page only."""
        import time
        start_time = time.time()

        try:
            cursor = self.connection.cursor()

            # Build paginated query
            offset = self._current_page * self._rows_per_page
            if self._show_all:
                paginated_sql = self._base_sql
            else:
                paginated_sql = self.adapter.add_pagination(self._base_sql, self._rows_per_page, offset)

            cursor.execute(paginated_sql)
            rows = cursor.fetchall()
            cursor.close()

            fetch_time = time.time() - start_time

            self.app.root.after(0, self._display_page_results, rows, fetch_time)
        except Exception as e:
            self.app.root.after(0, self._query_error, str(e), {})
        finally:
            self.app.root.after(0, self._set_running, False)

    def _display_page_results(self, rows, fetch_time):
        """Display page results."""
        self._all_rows = list(rows)
        self._refresh_display()
        self.app.statusbar.config(
            text=f"Page {self._current_page + 1} loaded ({len(rows)} rows, {fetch_time:.3f}s) from {self.conn_name}"
        )

    def _refresh_display(self):
        """Refresh the results display with current page data."""
        if not self._columns:
            return

        # Clear current display
        self.results_tree.delete(*self.results_tree.get_children())

        # Display all rows in _all_rows (which is the current page's data)
        for row in self._all_rows:
            clean_row = tuple(
                str(v).strip() if isinstance(v, str) else (str(v) if v is not None else "")
                for v in row
            )
            self.results_tree.insert("", tk.END, values=clean_row)

        # Calculate display range based on server-side pagination
        page_size = len(self._all_rows)
        if self._show_all:
            start_idx = 0
            end_idx = self._total_rows
        else:
            start_idx = self._current_page * self._rows_per_page
            end_idx = start_idx + page_size

        # Update pagination label and buttons
        self._update_pagination_ui(start_idx, end_idx, self._total_rows)

    def _update_pagination_ui(self, start_idx, end_idx, total_rows):
        """Update pagination label and button states."""
        if total_rows == 0:
            self.page_label.config(text="No results")
            self.first_btn.config(state=tk.DISABLED)
            self.prev_btn.config(state=tk.DISABLED)
            self.next_btn.config(state=tk.DISABLED)
            self.last_btn.config(state=tk.DISABLED)
        elif self._show_all:
            self.page_label.config(text=f"Showing all {total_rows:,} row(s)")
            self.first_btn.config(state=tk.DISABLED)
            self.prev_btn.config(state=tk.DISABLED)
            self.next_btn.config(state=tk.DISABLED)
            self.last_btn.config(state=tk.DISABLED)
        else:
            max_page = self._get_max_page()
            self.page_label.config(
                text=f"Showing {start_idx + 1:,}-{end_idx:,} of {total_rows:,} row(s)  "
                     f"(Page {self._current_page + 1} of {max_page + 1})"
            )

            # Update button states
            has_prev = self._current_page > 0
            has_next = self._current_page < max_page

            self.first_btn.config(state=tk.NORMAL if has_prev else tk.DISABLED)
            self.prev_btn.config(state=tk.NORMAL if has_prev else tk.DISABLED)
            self.next_btn.config(state=tk.NORMAL if has_next else tk.DISABLED)
            self.last_btn.config(state=tk.NORMAL if has_next else tk.DISABLED)

    def _set_running(self, running):
        """Update UI state for running/not running."""
        self._running = running
        if running:
            self.run_btn.config(state=tk.DISABLED)
            self.cancel_btn.config(state=tk.NORMAL)
            self.app.root.config(cursor="watch")
            self.app.statusbar.config(text=f"Running query on {self.conn_name}...")
        else:
            self.run_btn.config(state=tk.NORMAL)
            self.cancel_btn.config(state=tk.DISABLED)
            self.app.root.config(cursor="")

    def _cancel_query(self):
        """Cancel the running query."""
        if self._cursor:
            try:
                self._cursor.cancel()
                self.app.statusbar.config(text="Query cancelled")
            except Exception:
                pass

    def _run_query(self):
        if self._running:
            return

        sql = self.sql_text.get("1.0", tk.END).strip()
        if not sql:
            return

        # Clear previous results
        self._all_rows = []
        self._columns = []
        self._column_info = []
        self._current_page = 0
        self._base_sql = ""
        self._total_rows = 0
        self.results_tree.delete(*self.results_tree.get_children())
        for col in self.results_tree["columns"]:
            self.results_tree.heading(col, text="")

        # Clear fields tab
        self.fields_tree.delete(*self.fields_tree.get_children())

        # Clear results label
        self.page_label.config(text="")

        # Clear statistics tab
        self.stats_text.config(state=tk.NORMAL)
        self.stats_text.delete("1.0", tk.END)
        self.stats_text.insert("1.0", "Executing query...")
        self.stats_text.config(state=tk.DISABLED)

        self._set_running(True)

        # Run query in thread
        thread = threading.Thread(target=self._execute_query, args=(sql,), daemon=True)
        thread.start()

    def _add_row_limit(self, sql):
        """Add row limit to SELECT statements if not already limited."""
        sql_stripped = sql.strip()
        while sql_stripped.endswith(';'):
            sql_stripped = sql_stripped[:-1].strip()
        sql_upper = sql_stripped.upper()

        # Only modify SELECT statements
        if not sql_upper.startswith("SELECT"):
            return sql

        # Check if already has a row limit clause
        limit_keywords = ["FETCH FIRST", "FETCH NEXT", "LIMIT ", "OFFSET "]
        for keyword in limit_keywords:
            if keyword in sql_upper:
                return sql  # Already has a limit

        # Use rows_per_page as the limit (or default 1000 if showing all)
        if self._show_all:
            limit = 10000  # Cap at 10000 even for "show all"
        else:
            limit = self._rows_per_page

        # Use adapter to add pagination
        return self.adapter.add_pagination(sql_stripped, limit)

    def _has_limit_clause(self, sql_upper):
        """Check if SQL already has a row limit clause."""
        limit_keywords = ["FETCH FIRST", "FETCH NEXT", "LIMIT ", "OFFSET "]
        return any(keyword in sql_upper for keyword in limit_keywords)

    def _store_base_sql(self, sql, total_rows):
        """Store base SQL and total rows for pagination."""
        self._base_sql = sql
        self._total_rows = total_rows

    def _execute_query(self, sql):
        """Execute query in background thread."""
        import time
        start_time = time.time()
        stats_info = {
            "sql": sql,
            "start_time": start_time,
            "fetch_time": 0,
            "row_count": 0,
            "total_rows": 0,
            "column_count": 0,
            "explain_info": None
        }

        try:
            self._cursor = self.connection.cursor()

            # Try to get explain/query plan info before executing
            explain_info = self._get_explain_info(sql)
            stats_info["explain_info"] = explain_info

            # Strip whitespace and trailing semicolons (statement separators)
            sql_stripped = sql.strip()
            while sql_stripped.endswith(';'):
                sql_stripped = sql_stripped[:-1].strip()
            sql_upper = sql_stripped.upper()
            is_select = sql_upper.startswith("SELECT")

            # For SELECT statements, get total count first
            total_rows = 0
            if is_select and not self._has_limit_clause(sql_upper):
                try:
                    count_sql = self.adapter.get_count_sql(sql_stripped)
                    self._cursor.execute(count_sql)
                    total_rows = self._cursor.fetchone()[0]
                    stats_info["total_rows"] = total_rows
                except Exception:
                    # If count fails, we'll just show what we fetch
                    pass

            # Build paginated query for SELECT
            if is_select and not self._has_limit_clause(sql_upper):
                if self._show_all:
                    executed_sql = sql_stripped
                else:
                    offset = self._current_page * self._rows_per_page
                    executed_sql = self.adapter.add_pagination(sql_stripped, self._rows_per_page, offset)
                stats_info["limited"] = True
                # Store base SQL for pagination
                self.app.root.after(0, self._store_base_sql, sql_stripped, total_rows)
            else:
                executed_sql = sql_stripped
                stats_info["limited"] = False

            # Execute the query
            self._cursor.execute(executed_sql)
            exec_time = time.time()
            stats_info["exec_time"] = exec_time - start_time

            # Check if this was a SELECT
            if self._cursor.description:
                columns = [desc[0] for desc in self._cursor.description]
                column_info = list(self._cursor.description)
                stats_info["column_count"] = len(columns)

                rows = self._cursor.fetchall()
                fetch_time = time.time()
                stats_info["fetch_time"] = fetch_time - exec_time
                stats_info["row_count"] = len(rows)

                # If we couldn't get count earlier, use fetched row count
                if total_rows == 0:
                    stats_info["total_rows"] = len(rows)

                # Update UI in main thread
                self.app.root.after(0, self._display_results, columns, rows, column_info, stats_info)
            else:
                self.connection.commit()
                stats_info["exec_time"] = time.time() - start_time
                self.app.root.after(0, self._query_done, "Statement executed successfully", stats_info)

            self._cursor.close()
        except Exception as e:
            stats_info["error"] = str(e)
            self.app.root.after(0, self._query_error, str(e), stats_info)
        finally:
            self._cursor = None
            self.app.root.after(0, self._set_running, False)

    def _get_explain_info(self, sql):
        """Try to get query explain/plan information."""
        explain_data = []

        # Only try to explain SELECT statements
        sql_upper = sql.strip().upper()
        if not sql_upper.startswith("SELECT"):
            return None

        # Explain is currently only implemented for IBM i
        if self.adapter.db_type != "ibmi":
            return None

        try:
            # Use Visual Explain via QSYS2
            explain_cursor = self.connection.cursor()

            # First, try to use EXPLAIN
            try:
                # Set up explain tables
                explain_cursor.execute("CALL QSYS2.OVERRIDE_QAQQINI(1, '', '')")
            except Exception:
                pass

            # Try to get access plan info using SYSIBM.SQLSTATISTICS or explain
            try:
                # This queries index usage from the system catalog for tables in the query
                # Extract table names from a simple SELECT (basic parsing)
                tables = self._extract_tables_from_sql(sql)
                if tables:
                    for table in tables[:5]:  # Limit to 5 tables
                        try:
                            idx_cursor = self.connection.cursor()
                            idx_cursor.execute("""
                                SELECT INDEX_NAME, COLUMN_NAME, INDEX_TYPE, IS_UNIQUE
                                FROM QSYS2.SYSINDEXES I
                                JOIN QSYS2.SYSKEYS K ON I.INDEX_NAME = K.INDEX_NAME
                                    AND I.INDEX_SCHEMA = K.INDEX_SCHEMA
                                WHERE I.TABLE_NAME = ?
                                ORDER BY I.INDEX_NAME, K.ORDINAL_POSITION
                                FETCH FIRST 20 ROWS ONLY
                            """, (table.upper(),))
                            indexes = idx_cursor.fetchall()
                            if indexes:
                                explain_data.append(f"\nIndexes on {table}:")
                                current_idx = None
                                for row in indexes:
                                    idx_name, col_name, idx_type, is_unique = row
                                    if idx_name != current_idx:
                                        unique_str = "UNIQUE " if is_unique == 'Y' else ""
                                        explain_data.append(f"  {unique_str}{idx_name} ({idx_type})")
                                        current_idx = idx_name
                                    explain_data.append(f"    - {col_name}")
                            idx_cursor.close()
                        except Exception:
                            pass
            except Exception:
                pass

            explain_cursor.close()
        except Exception:
            pass

        return "\n".join(explain_data) if explain_data else None

    def _extract_tables_from_sql(self, sql):
        """Extract table names from SQL (basic parsing). Returns schema.table format."""
        import re
        tables = []
        # Simple regex to find FROM and JOIN clauses
        # This is a basic implementation - won't catch all cases

        # Find tables after FROM (handles schema.table and library/table formats)
        from_match = re.search(r'\bFROM\s+([A-Za-z0-9_.]+(?:/[A-Za-z0-9_]+)?)', sql, re.IGNORECASE)
        if from_match:
            table = from_match.group(1)
            # Convert library/table to schema.table format
            if '/' in table:
                parts = table.split('/')
                table = f"{parts[0]}.{parts[1]}"
            tables.append(table)

        # Find tables after JOIN
        join_matches = re.findall(r'\bJOIN\s+([A-Za-z0-9_.]+(?:/[A-Za-z0-9_]+)?)', sql, re.IGNORECASE)
        for match in join_matches:
            table = match
            if '/' in table:
                parts = table.split('/')
                table = f"{parts[0]}.{parts[1]}"
            tables.append(table)

        return tables

    def _calculate_column_width(self, col_name, col_info):
        """Calculate appropriate column width based on field metadata and font size."""
        scale = self.app.font_size / 10.0
        # Base width on column name (approx 8 pixels per char at font size 10)
        name_width = len(col_name) * 8 + 20

        if not col_info:
            return int(max(name_width, 100) * scale)

        # Use adapter to get the best display size for this driver
        char_count = self.adapter.get_column_display_size(col_info)

        # Calculate data width (approx 8 pixels per character at font size 10)
        if char_count <= 5:
            data_width = 60
        elif char_count <= 10:
            data_width = 80
        elif char_count <= 20:
            data_width = 130
        elif char_count <= 50:
            data_width = 200
        else:
            data_width = min(char_count * 6, 300)

        # Return the larger of name width or data width, scaled by font size
        base_width = max(60, min(max(name_width, data_width), 350))
        return int(base_width * scale)

    def _get_column_anchor(self, col_info):
        """Determine text alignment based on column type."""
        if not col_info:
            return "w"  # left align by default

        type_code = col_info[1]

        # Use adapter to check for numeric type
        if self.adapter.is_numeric_type(type_code):
            return "e"  # right align numbers

        # Also check type name string for common numeric types (fallback)
        type_name = getattr(type_code, '__name__', '').lower() if type_code else ''
        if type_name in ('int', 'float', 'decimal', 'numeric', 'integer',
                         'smallint', 'bigint', 'real', 'double'):
            return "e"

        return "w"  # left align everything else

    def _sort_by_column(self, col):
        """Sort results by the clicked column."""
        if not self._all_rows or not self._columns:
            return

        # Toggle direction if same column, otherwise sort ascending
        if self._sort_column == col:
            self._sort_reverse = not self._sort_reverse
        else:
            self._sort_column = col
            self._sort_reverse = False

        # Get column index
        try:
            col_idx = self._columns.index(col)
        except ValueError:
            return

        # Sort the data
        def sort_key(row):
            val = row[col_idx]
            # Handle None values - sort them last
            if val is None:
                return (1, "")
            # Try numeric comparison
            if isinstance(val, (int, float)):
                return (0, val)
            try:
                from decimal import Decimal
                if isinstance(val, Decimal):
                    return (0, float(val))
            except Exception:
                pass
            # String comparison
            return (0, str(val).strip().lower())

        self._all_rows.sort(key=sort_key, reverse=self._sort_reverse)

        # Update column heading to show sort indicator
        for c in self._columns:
            if c == col:
                indicator = " ▼" if self._sort_reverse else " ▲"
                self.results_tree.heading(c, text=c + indicator)
            else:
                self.results_tree.heading(c, text=c)

        # Refresh display
        self._refresh_display()

    def _display_results(self, columns, rows, column_info, stats_info):
        """Display results in the treeview (called from main thread)."""
        # Store all results
        self._columns = columns
        self._all_rows = list(rows)
        self._column_info = column_info
        self._current_page = 0

        # Setup columns in results tree
        self._sort_column = None
        self._sort_reverse = False
        self._results_base_widths = {}  # Reset base widths for new query
        scale = self.app.font_size / 10.0
        self.results_tree["columns"] = columns
        for i, col in enumerate(columns):
            self.results_tree.heading(col, text=col, command=lambda c=col: self._sort_by_column(c))
            col_info = column_info[i] if i < len(column_info) else None
            width = self._calculate_column_width(col, col_info)
            # Store base width (unscaled) for later rescaling
            self._results_base_widths[col] = int(width / scale) if scale else width
            anchor = self._get_column_anchor(col_info)
            self.results_tree.column(col, width=width, minwidth=40, stretch=False, anchor=anchor)

        # Display current page
        self._refresh_display()

        # Update fields tab
        self._update_fields_tab(column_info)

        # Update statistics tab
        self._update_statistics_tab(stats_info)

        self.app.statusbar.config(text=f"{len(rows):,} row(s) returned from {self.conn_name}")

    def _update_fields_tab(self, column_info):
        """Update the Fields tab with column metadata."""
        self.fields_tree.delete(*self.fields_tree.get_children())

        # Try to get table info for columns
        table_map = self._get_column_tables([desc[0] for desc in column_info])

        # cursor.description format: (name, type_code, display_size, internal_size, precision, scale, null_ok)
        for desc in column_info:
            name = desc[0]
            type_code = desc[1]
            display_size = desc[2] if desc[2] else ""
            precision = desc[4] if desc[4] else ""
            scale = desc[5] if desc[5] else ""
            nullable = "Yes" if desc[6] else "No"

            # Get table name for this column
            table_name = table_map.get(name.upper(), "")

            # Try to get type name
            type_name = getattr(type_code, '__name__', str(type_code)) if type_code else "UNKNOWN"

            self.fields_tree.insert("", tk.END, values=(
                table_name, name, type_name, display_size, precision, scale, nullable
            ))

    def _get_column_tables(self, column_names):
        """Try to determine which table each column belongs to."""
        table_map = {}

        if not self._base_sql:
            return table_map

        # Extract table names from SQL
        tables = self._extract_tables_from_sql(self._base_sql)

        if not tables:
            return table_map

        # If only one table, all columns belong to it
        if len(tables) == 1:
            table_name = tables[0].split('.')[-1]  # Get just table name, not schema
            for col in column_names:
                table_map[col.upper()] = table_name.upper()
            return table_map

        # Multiple tables - try to look up each column in system catalog
        columns_query = self.adapter.get_columns_query(tables)
        if not columns_query:
            return table_map

        try:
            cursor = self.connection.cursor()
            cursor.execute(columns_query)
            rows = cursor.fetchall()
            cursor.close()

            # Build lookup of column -> table from results
            col_to_table = {}
            for row in rows:
                # Row format: (schema, table, column, type, length, scale)
                table_name = row[1]
                col_name = row[2]
                col_to_table[col_name.upper()] = table_name.upper()

            # Map requested columns
            for col in column_names:
                col_upper = col.upper()
                if col_upper in col_to_table:
                    table_map[col_upper] = col_to_table[col_upper]
        except Exception:
            pass

        return table_map

    def _update_statistics_tab(self, stats_info):
        """Update the Statistics tab with query execution info."""
        self.stats_text.config(state=tk.NORMAL)
        self.stats_text.delete("1.0", tk.END)

        lines = []
        lines.append("=" * 60)
        lines.append("QUERY EXECUTION STATISTICS")
        lines.append("=" * 60)
        lines.append("")

        # Timing info
        lines.append("TIMING:")
        lines.append(f"  Execution time: {stats_info.get('exec_time', 0):.3f} seconds")
        if stats_info.get('fetch_time'):
            lines.append(f"  Fetch time:     {stats_info.get('fetch_time', 0):.3f} seconds")
            total_time = stats_info.get('exec_time', 0) + stats_info.get('fetch_time', 0)
            lines.append(f"  Total time:     {total_time:.3f} seconds")
        lines.append("")

        # Result info
        lines.append("RESULTS:")
        row_count = stats_info.get('row_count', 0)
        total_rows = stats_info.get('total_rows', row_count)
        if stats_info.get('limited') and total_rows > row_count:
            lines.append(f"  Rows fetched:   {row_count:,}")
            lines.append(f"  Total rows:     {total_rows:,}")
        else:
            lines.append(f"  Rows returned:  {row_count:,}")
        lines.append(f"  Columns:        {stats_info.get('column_count', 0)}")
        lines.append("")

        # Query
        lines.append("QUERY:")
        lines.append("-" * 40)
        sql = stats_info.get('sql', '')
        lines.append(sql[:500] + ('...' if len(sql) > 500 else ''))
        lines.append("")

        # Explain/Index info
        if stats_info.get('explain_info'):
            lines.append("INDEX INFORMATION:")
            lines.append("-" * 40)
            lines.append(stats_info['explain_info'])
            lines.append("")

        self.stats_text.insert("1.0", "\n".join(lines))
        self.stats_text.config(state=tk.DISABLED)

    def _query_done(self, message, stats_info=None):
        """Called when non-SELECT query completes."""
        self._all_rows = []
        self._columns = []
        self._update_pagination_ui(0, 0, 0)

        if stats_info:
            self._update_statistics_tab(stats_info)

        self.app.statusbar.config(text=f"{message} on {self.conn_name}")

    def _query_error(self, error, stats_info=None):
        """Called when query fails."""
        if stats_info:
            stats_info['error'] = error
            self.stats_text.config(state=tk.NORMAL)
            self.stats_text.delete("1.0", tk.END)
            self.stats_text.insert("1.0", f"ERROR:\n{error}\n\nQuery:\n{stats_info.get('sql', '')}")
            self.stats_text.config(state=tk.DISABLED)

        messagebox.showerror("SQL Error", error)

    def _save_query(self):
        sql = self.sql_text.get("1.0", tk.END).strip()
        if not sql:
            messagebox.showwarning("Empty", "No SQL to save.")
            return

        name = simpledialog.askstring("Save Query", "Enter a name for this query:")
        if name:
            self.app.db.save_query(name, sql, self.conn_name, self.adapter.db_type)
            messagebox.showinfo("Saved", f"Query '{name}' saved for {self.adapter.db_type}.")

    def _load_query(self):
        db_type = self.adapter.db_type
        all_queries = self.app.db.get_saved_queries(db_type)

        # Query management dialog
        load_win = tk.Toplevel(self.frame)
        load_win.title(f"Manage Queries ({db_type})")
        load_win.geometry("450x450")
        load_win.transient(self.frame.winfo_toplevel())
        load_win.grab_set()
        load_win.bind("<Escape>", lambda e: load_win.destroy())

        # Apply theme to dialog
        is_dark = self.app.dark_mode_var.get()
        if is_dark:
            load_win.configure(bg="#2b2b2b")

        # Filter frame
        filter_frame = ttk.Frame(load_win)
        filter_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        ttk.Label(filter_frame, text="Filter:").pack(side=tk.LEFT)
        filter_var = tk.StringVar()
        filter_entry = ttk.Entry(filter_frame, textvariable=filter_var)
        filter_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))

        # Header
        ttk.Label(load_win, text=f"Queries for {db_type}:", font=("TkDefaultFont", 9, "bold")).pack(anchor=tk.W, padx=10)

        # Listbox with scrollbar
        list_frame = ttk.Frame(load_win)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)

        # Apply theme colors
        is_dark = self.app.dark_mode_var.get()
        if is_dark:
            listbox.configure(bg="#313335", fg="#a9b7c6",
                            selectbackground="#214283", selectforeground="#a9b7c6")

        # Track filtered queries
        filtered_queries = []

        def refresh_list(filter_text=""):
            nonlocal filtered_queries
            listbox.delete(0, tk.END)
            filtered_queries = []
            for q in all_queries:
                name = q["name"]
                if filter_text.lower() in name.lower():
                    suffix = "" if q.get("db_type") else " (any)"
                    listbox.insert(tk.END, f"{name}{suffix}")
                    filtered_queries.append(q)
            if not filtered_queries and not all_queries:
                listbox.insert(tk.END, "(no saved queries)")

        def on_filter(*args):
            refresh_list(filter_var.get())

        filter_var.trace("w", on_filter)
        refresh_list()

        def get_selected():
            sel = listbox.curselection()
            if sel and filtered_queries:
                return filtered_queries[sel[0]]
            return None

        def on_load():
            query = get_selected()
            if query:
                self.sql_text.delete("1.0", tk.END)
                self.sql_text.insert("1.0", query["sql"])
                load_win.destroy()

        def on_delete():
            query = get_selected()
            if query:
                if messagebox.askyesno("Delete Query", f"Delete '{query['name']}'?", parent=load_win):
                    self.app.db.delete_query(query["id"])
                    all_queries.remove(query)
                    refresh_list(filter_var.get())

        def on_export():
            if not all_queries:
                messagebox.showinfo("No Queries", "No queries to export.", parent=load_win)
                return
            filepath = filedialog.asksaveasfilename(
                parent=load_win,
                title="Export Queries",
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
                initialfile=f"queries_{db_type}.json"
            )
            if filepath:
                import json
                export_data = [{"name": q["name"], "sql": q["sql"], "db_type": q.get("db_type") or db_type}
                              for q in all_queries]
                with open(filepath, "w") as f:
                    json.dump(export_data, f, indent=2)
                messagebox.showinfo("Exported", f"Exported {len(export_data)} queries.", parent=load_win)

        def on_import():
            nonlocal all_queries
            filepath = filedialog.askopenfilename(
                parent=load_win,
                title="Import Queries",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )
            if filepath:
                import json
                try:
                    with open(filepath, "r") as f:
                        import_data = json.load(f)
                    count = 0
                    for q in import_data:
                        name = q.get("name")
                        sql = q.get("sql")
                        q_db_type = q.get("db_type", db_type)
                        if name and sql:
                            # Only import queries matching current db_type or untyped
                            if q_db_type == db_type or q_db_type is None:
                                self.app.db.save_query(name, sql, None, db_type)
                                count += 1
                    # Refresh query list
                    all_queries = self.app.db.get_saved_queries(db_type)
                    refresh_list(filter_var.get())
                    messagebox.showinfo("Imported", f"Imported {count} queries.", parent=load_win)
                except Exception as e:
                    messagebox.showerror("Import Error", str(e), parent=load_win)

        listbox.bind("<Double-1>", lambda e: on_load())

        # Button frame
        btn_frame = ttk.Frame(load_win)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Button(btn_frame, text="Load", command=on_load).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Delete", command=on_delete).pack(side=tk.LEFT, padx=2)
        ttk.Separator(btn_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)
        ttk.Button(btn_frame, text="Import", command=on_import).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Export", command=on_export).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Close", command=load_win.destroy).pack(side=tk.RIGHT, padx=2)

        filter_entry.focus()

    def _clear(self):
        self.sql_text.delete("1.0", tk.END)
        self._all_rows = []
        self._columns = []
        self._column_info = []
        self._current_page = 0
        self.results_tree.delete(*self.results_tree.get_children())
        self.fields_tree.delete(*self.fields_tree.get_children())
        self._update_pagination_ui(0, 0, 0)

        self.stats_text.config(state=tk.NORMAL)
        self.stats_text.delete("1.0", tk.END)
        self.stats_text.insert("1.0", "Execute a query to see statistics.")
        self.stats_text.config(state=tk.DISABLED)

    def _copy_to_clipboard(self):
        """Copy current results to clipboard as tab-separated values."""
        if not self._all_rows:
            messagebox.showwarning("No Data", "No results to copy.")
            return

        lines = []

        # Add headers
        lines.append("\t".join(self._columns))

        # Add data rows
        for row in self._all_rows:
            clean_row = []
            for value in row:
                if isinstance(value, str):
                    clean_row.append(value.strip())
                elif value is None:
                    clean_row.append("")
                else:
                    clean_row.append(str(value))
            lines.append("\t".join(clean_row))

        # Copy to clipboard
        text = "\n".join(lines)
        self.app.root.clipboard_clear()
        self.app.root.clipboard_append(text)
        self.app.statusbar.config(text=f"Copied {len(self._all_rows)} rows to clipboard")

    def _save_to_excel(self):
        """Save current results to Excel file."""
        if not self._all_rows:
            messagebox.showwarning("No Data", "No results to save.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="query_results.xlsx"
        )
        if not file_path:
            return

        try:
            import openpyxl
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"

            # Write headers
            for col_idx, col_name in enumerate(self._columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)

            # Write data
            for row_idx, row in enumerate(self._all_rows, 2):
                for col_idx, value in enumerate(row, 1):
                    # Strip strings, convert None to empty
                    if isinstance(value, str):
                        value = value.strip()
                    elif value is None:
                        value = ""
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for col_idx, col_name in enumerate(self._columns, 1):
                max_len = len(str(col_name))
                for row in ws.iter_rows(min_row=2, max_row=min(100, len(self._all_rows) + 1), min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

            wb.save(file_path)
            self.app.statusbar.config(text=f"Saved to {file_path}")
            messagebox.showinfo("Saved", f"Results saved to:\n{file_path}")
        except ImportError:
            messagebox.showerror("Error", "openpyxl not installed. Run: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Could not save Excel file: {e}")

    def _save_to_csv(self):
        """Save current results to CSV file."""
        if not self._all_rows:
            messagebox.showwarning("No Data", "No results to save.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile="query_results.csv"
        )
        if not file_path:
            return

        try:
            import csv

            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)

                # Write headers
                writer.writerow(self._columns)

                # Write data
                for row in self._all_rows:
                    clean_row = []
                    for value in row:
                        if isinstance(value, str):
                            clean_row.append(value.strip())
                        elif value is None:
                            clean_row.append("")
                        else:
                            clean_row.append(value)
                    writer.writerow(clean_row)

            self.app.statusbar.config(text=f"Saved to {file_path}")
            messagebox.showinfo("Saved", f"Results saved to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not save CSV file: {e}")

    def _save_to_json(self):
        """Save current results to JSON file."""
        if not self._all_rows:
            messagebox.showwarning("No Data", "No results to save.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            initialfile="query_results.json"
        )
        if not file_path:
            return

        try:
            import json
            from datetime import date, datetime
            from decimal import Decimal

            def json_serializer(obj):
                if isinstance(obj, (datetime, date)):
                    return obj.isoformat()
                elif isinstance(obj, Decimal):
                    return float(obj)
                elif isinstance(obj, bytes):
                    return obj.decode('utf-8', errors='replace')
                return str(obj)

            # Convert rows to list of dicts
            data = []
            for row in self._all_rows:
                record = {}
                for i, col in enumerate(self._columns):
                    value = row[i]
                    if isinstance(value, str):
                        value = value.strip()
                    record[col] = value
                data.append(record)

            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, default=json_serializer)

            self.app.statusbar.config(text=f"Saved to {file_path}")
            messagebox.showinfo("Saved", f"Results saved to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not save JSON file: {e}")

    def get_sql(self):
        """Get current SQL text."""
        return self.sql_text.get("1.0", tk.END).strip()

    def set_sql(self, sql):
        """Set SQL text."""
        self.sql_text.delete("1.0", tk.END)
        self.sql_text.insert("1.0", sql)
        # Apply syntax highlighting
        self._highlight_sql()

    def scale_columns(self, scale):
        """Scale column widths based on font size ratio (scale = font_size / 10)."""
        # Scale fields_tree columns
        for col, base_w in self._fields_base_widths.items():
            try:
                self.fields_tree.column(col, width=int(base_w * scale))
            except Exception:
                pass

        # Scale results_tree columns if they exist
        if hasattr(self, '_results_base_widths') and self._results_base_widths:
            for col, base_w in self._results_base_widths.items():
                try:
                    self.results_tree.column(col, width=int(base_w * scale))
                except Exception:
                    pass

    def _on_results_double_click(self, event):
        """Handle double-click on results treeview - autosize column or open record viewer."""
        region = self.results_tree.identify_region(event.x, event.y)

        if region == "separator":
            # Double-click on header separator - autosize the column to the LEFT
            col_id = self.results_tree.identify_column(event.x)
            if col_id:
                # col_id is like "#1", "#2", etc. - get the column to the LEFT of separator
                col_num = int(col_id.replace("#", ""))
                if col_num > 0:
                    columns = self.results_tree["columns"]
                    if col_num <= len(columns):
                        col_name = columns[col_num - 1]
                        self._autosize_column(col_name)
            return "break"  # Prevent other handlers
        elif region == "cell":
            # Double-click on data cell - open record viewer
            self._open_record_viewer(event)

    def _autosize_column(self, col_name):
        """Auto-size a column to fit the maximum content width."""
        if not self._all_rows or not self._columns:
            return

        try:
            col_index = list(self._columns).index(col_name)
        except ValueError:
            return

        # Calculate max width needed
        # Start with header width
        import tkinter.font as tkfont
        font = tkfont.nametofont("TkDefaultFont")
        max_width = font.measure(str(col_name)) + 20  # Add padding for header

        # Check all rows (use displayed rows for performance with large datasets)
        rows_to_check = self._all_rows[:1000]  # Limit to first 1000 rows for performance
        for row in rows_to_check:
            if col_index < len(row):
                value = row[col_index]
                if value is not None:
                    text_width = font.measure(str(value)) + 20  # Add padding
                    max_width = max(max_width, text_width)

        # Clamp to reasonable bounds
        max_width = max(50, min(max_width, 600))  # Min 50, max 600 pixels

        self.results_tree.column(col_name, width=max_width)

    def _open_record_viewer(self, event=None):
        """Open record viewer dialog for the selected row."""
        if not self._all_rows or not self._columns:
            return

        # Get selected item
        selection = self.results_tree.selection()
        if not selection:
            return

        # Calculate row index from position in treeview + page offset
        selected_item = selection[0]
        all_items = self.results_tree.get_children()
        try:
            row_in_page = list(all_items).index(selected_item)
            row_index = self._current_page * self._rows_per_page + row_in_page
        except ValueError:
            row_index = 0

        # Open the dialog with callback to sync selection
        RecordViewerDialog(
            self.frame.winfo_toplevel(),
            self._columns,
            self._all_rows,
            row_index,
            self.app,
            on_navigate=self._sync_results_selection
        )

    def _sync_results_selection(self, row_index):
        """Sync the results tree selection with the record viewer."""
        # Only update selection if row is on current page
        page = row_index // self._rows_per_page
        if page != self._current_page:
            return  # Don't change pages, just skip

        all_items = self.results_tree.get_children()
        row_in_page = row_index % self._rows_per_page
        if row_in_page < len(all_items):
            item = all_items[row_in_page]
            self.results_tree.selection_set(item)
            self.results_tree.see(item)

    # ===== SQL Editor Context Menu and Search =====

    def _create_sql_context_menu(self):
        """Create right-click context menu for SQL editor."""
        self.sql_context_menu = tk.Menu(self.sql_text, tearoff=0)
        self.sql_context_menu.add_command(label="Select All", command=self._select_all, accelerator="Ctrl+A")
        self.sql_context_menu.add_separator()
        self.sql_context_menu.add_command(label="Cut", command=self._sql_cut, accelerator="Ctrl+X")
        self.sql_context_menu.add_command(label="Copy", command=self._sql_copy, accelerator="Ctrl+C")
        self.sql_context_menu.add_command(label="Paste", command=self._sql_paste, accelerator="Ctrl+V")
        self.sql_context_menu.add_separator()
        self.sql_context_menu.add_command(label="Find...", command=self._show_sql_search_dialog, accelerator="Ctrl+F")

        self.sql_text.bind("<Button-3>", self._show_sql_context_menu)
        self.sql_text.bind("<Control-f>", lambda e: self._show_sql_search_dialog())
        self.sql_text.bind("<Control-F>", lambda e: self._show_sql_search_dialog())

    def _show_sql_context_menu(self, event):
        """Show context menu at mouse position."""
        try:
            self.sql_context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.sql_context_menu.grab_release()

    def _sql_cut(self):
        """Cut selected text to clipboard."""
        try:
            selected = self.sql_text.get("sel.first", "sel.last")
            self.app.root.clipboard_clear()
            self.app.root.clipboard_append(selected)
            self.sql_text.delete("sel.first", "sel.last")
        except tk.TclError:
            pass  # No selection

    def _sql_copy(self):
        """Copy selected text to clipboard."""
        try:
            selected = self.sql_text.get("sel.first", "sel.last")
            self.app.root.clipboard_clear()
            self.app.root.clipboard_append(selected)
        except tk.TclError:
            pass  # No selection

    def _sql_paste(self):
        """Paste text from clipboard."""
        try:
            # Delete selection if any
            try:
                self.sql_text.delete("sel.first", "sel.last")
            except tk.TclError:
                pass
            # Get clipboard content and insert
            text = self.app.root.clipboard_get()
            self.sql_text.insert("insert", text)
        except tk.TclError:
            pass  # Nothing in clipboard

    def _show_sql_search_dialog(self):
        """Show search dialog for SQL editor."""
        # Create a simple search bar at the top of the SQL frame
        if hasattr(self, '_sql_search_frame') and self._sql_search_frame.winfo_exists():
            # Already visible, focus the entry
            self._sql_search_entry.focus_set()
            self._sql_search_entry.select_range(0, tk.END)
            return

        # Get the SQL frame (parent of sql_text)
        sql_frame = self.sql_text.master

        # Create search frame
        self._sql_search_frame = ttk.Frame(sql_frame)
        self._sql_search_frame.pack(side=tk.TOP, fill=tk.X, before=self.sql_text)

        ttk.Label(self._sql_search_frame, text="Find:").pack(side=tk.LEFT, padx=(5, 2))

        self._sql_search_var = tk.StringVar()
        self._sql_search_entry = ttk.Entry(self._sql_search_frame, textvariable=self._sql_search_var, width=25)
        self._sql_search_entry.pack(side=tk.LEFT, padx=2)
        self._sql_search_entry.bind("<Return>", lambda e: self._sql_search_next())
        self._sql_search_entry.bind("<Shift-Return>", lambda e: self._sql_search_prev())
        self._sql_search_entry.bind("<Escape>", lambda e: self._close_sql_search())

        ttk.Button(self._sql_search_frame, text="<", width=2, command=self._sql_search_prev).pack(side=tk.LEFT, padx=1)
        ttk.Button(self._sql_search_frame, text=">", width=2, command=self._sql_search_next).pack(side=tk.LEFT, padx=1)

        self._sql_search_label = ttk.Label(self._sql_search_frame, text="")
        self._sql_search_label.pack(side=tk.LEFT, padx=10)

        ttk.Button(self._sql_search_frame, text="X", width=2, command=self._close_sql_search).pack(side=tk.RIGHT, padx=2)

        self._sql_search_entry.focus_set()

    def _close_sql_search(self):
        """Close the SQL search bar."""
        if hasattr(self, '_sql_search_frame') and self._sql_search_frame.winfo_exists():
            self._sql_search_frame.destroy()
            # Clear highlights
            self.sql_text.tag_remove("search_highlight", "1.0", tk.END)
            self.sql_text.tag_remove("search_current", "1.0", tk.END)
            self._sql_search_matches = []
            self._sql_search_index = -1
            self.sql_text.focus_set()

    def _sql_search_next(self):
        """Find next occurrence in SQL editor."""
        if not hasattr(self, '_sql_search_var'):
            return
        search_term = self._sql_search_var.get()
        if not search_term:
            return

        # If search term changed, find all matches
        if not self._sql_search_matches or self._sql_last_search != search_term:
            self._sql_find_all_matches(search_term)
            self._sql_last_search = search_term

        if not self._sql_search_matches:
            self._sql_search_label.config(text="Not found")
            return

        # Move to next match
        self._sql_search_index = (self._sql_search_index + 1) % len(self._sql_search_matches)
        self._sql_highlight_current_match()

    def _sql_search_prev(self):
        """Find previous occurrence in SQL editor."""
        if not hasattr(self, '_sql_search_var'):
            return
        search_term = self._sql_search_var.get()
        if not search_term:
            return

        # If search term changed, find all matches
        if not self._sql_search_matches or self._sql_last_search != search_term:
            self._sql_find_all_matches(search_term)
            self._sql_last_search = search_term

        if not self._sql_search_matches:
            self._sql_search_label.config(text="Not found")
            return

        # Move to previous match
        self._sql_search_index = (self._sql_search_index - 1) % len(self._sql_search_matches)
        self._sql_highlight_current_match()

    def _sql_find_all_matches(self, search_term):
        """Find all matches in SQL editor."""
        self._sql_search_matches = []
        self._sql_search_index = -1

        # Remove existing highlights
        self.sql_text.tag_remove("search_highlight", "1.0", tk.END)
        self.sql_text.tag_remove("search_current", "1.0", tk.END)

        if not search_term:
            return

        # Search through text (case insensitive)
        start_pos = "1.0"
        while True:
            pos = self.sql_text.search(search_term, start_pos, stopindex=tk.END, nocase=True)
            if not pos:
                break

            end_pos = f"{pos}+{len(search_term)}c"
            self._sql_search_matches.append((pos, end_pos))
            self.sql_text.tag_add("search_highlight", pos, end_pos)
            start_pos = end_pos

    def _sql_highlight_current_match(self):
        """Highlight the current match in SQL editor."""
        if not self._sql_search_matches or self._sql_search_index < 0:
            return

        # Remove previous current highlight
        self.sql_text.tag_remove("search_current", "1.0", tk.END)

        # Apply current highlight
        pos, end_pos = self._sql_search_matches[self._sql_search_index]
        self.sql_text.tag_add("search_current", pos, end_pos)

        # Scroll to match
        self.sql_text.see(pos)
        self.sql_text.mark_set("insert", pos)

        # Update label
        if hasattr(self, '_sql_search_label'):
            self._sql_search_label.config(
                text=f"{self._sql_search_index + 1} of {len(self._sql_search_matches)}"
            )

    # ===== Results Search =====

    def _results_search_next(self):
        """Find next occurrence in results."""
        search_term = self.results_search_var.get()
        if not search_term:
            return

        # If search term changed, find all matches
        if not self._results_search_matches or self._results_last_search != search_term:
            self._results_find_all_matches(search_term)
            self._results_last_search = search_term

        if not self._results_search_matches:
            self.app.statusbar.config(text=f"'{search_term}' not found in results")
            return

        # Move to next match
        self._results_search_index = (self._results_search_index + 1) % len(self._results_search_matches)
        self._results_highlight_current_match()

    def _results_search_prev(self):
        """Find previous occurrence in results."""
        search_term = self.results_search_var.get()
        if not search_term:
            return

        # If search term changed, find all matches
        if not self._results_search_matches or self._results_last_search != search_term:
            self._results_find_all_matches(search_term)
            self._results_last_search = search_term

        if not self._results_search_matches:
            self.app.statusbar.config(text=f"'{search_term}' not found in results")
            return

        # Move to previous match
        self._results_search_index = (self._results_search_index - 1) % len(self._results_search_matches)
        self._results_highlight_current_match()

    def _results_find_all_matches(self, search_term):
        """Find all matches in results treeview."""
        self._results_search_matches = []
        self._results_search_index = -1

        if not search_term or not self._all_rows:
            return

        search_lower = search_term.lower()

        # Search through all visible items in treeview
        for item_id in self.results_tree.get_children():
            values = self.results_tree.item(item_id, 'values')
            for col_idx, value in enumerate(values):
                if search_lower in str(value).lower():
                    self._results_search_matches.append((item_id, col_idx))

    def _results_highlight_current_match(self):
        """Highlight the current match in results."""
        if not self._results_search_matches or self._results_search_index < 0:
            return

        item_id, col_idx = self._results_search_matches[self._results_search_index]

        # Select and scroll to the item
        self.results_tree.selection_set(item_id)
        self.results_tree.see(item_id)
        self.results_tree.focus(item_id)

        # Update status
        self.app.statusbar.config(
            text=f"Match {self._results_search_index + 1} of {len(self._results_search_matches)} "
                 f"(column: {self._columns[col_idx] if col_idx < len(self._columns) else '?'})"
        )

    def apply_theme(self):
        """Apply current theme to SQL tab components."""
        is_dark = self.app.dark_mode_var.get()

        if is_dark:
            menu_bg = "#2b2b2b"
            menu_fg = "#a9b7c6"
            select_bg = "#214283"
            search_highlight = "#614900"
            search_current = "#8B4000"
            highlight_fg = "#ffffff"
        else:
            menu_bg = "#f0f0f0"
            menu_fg = "#000000"
            select_bg = "#0078d4"
            search_highlight = "#FFFF00"
            search_current = "#FF8C00"
            highlight_fg = "#000000"

        # Update context menu
        if hasattr(self, 'sql_context_menu'):
            self.sql_context_menu.configure(bg=menu_bg, fg=menu_fg,
                                           activebackground=select_bg, activeforeground=menu_fg)

        # Update search highlight tags
        self.sql_text.tag_configure("search_highlight", background=search_highlight, foreground=highlight_fg)
        self.sql_text.tag_configure("search_current", background=search_current, foreground=highlight_fg)

        # Update syntax highlighting for theme
        self._setup_syntax_highlighting()
        self._highlight_sql()


class RecordViewerDialog:
    """Dialog to view a single record with navigation."""

    def __init__(self, parent, columns, rows, initial_index, app, on_navigate=None):
        self.columns = columns
        self.rows = rows
        self.current_index = initial_index
        self.app = app
        self.on_navigate = on_navigate

        self.top = tk.Toplevel(parent)
        self.top.title("Record Viewer")
        self.top.transient(parent)

        # Size based on number of fields - larger for more fields
        num_fields = len(columns)
        width = 700
        height = min(max(400, num_fields * 28 + 100), 800)

        # Size and position
        self.top.geometry(f"{width}x{height}")
        self.top.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - width) // 2
        y = parent.winfo_y() + (parent.winfo_height() - height) // 2
        self.top.geometry(f"+{x}+{y}")

        # Apply theme
        self._apply_theme()

        self._create_widgets()
        self._display_record()

        # Key bindings
        self.top.bind("<Left>", lambda e: self._prev_record())
        self.top.bind("<Right>", lambda e: self._next_record())
        self.top.bind("<Home>", lambda e: self._first_record())
        self.top.bind("<End>", lambda e: self._last_record())
        self.top.bind("<Escape>", lambda e: self.top.destroy())

    def _apply_theme(self):
        """Apply dark/light theme."""
        is_dark = self.app.dark_mode_var.get()
        if is_dark:
            self.bg = "#2b2b2b"
            self.fg = "#a9b7c6"
            self.text_bg = "#313335"
            self.select_bg = "#214283"
        else:
            self.bg = "#f0f0f0"
            self.fg = "#000000"
            self.text_bg = "#ffffff"
            self.select_bg = "#0078d4"

        self.top.configure(bg=self.bg)

    def _create_widgets(self):
        """Create dialog widgets."""
        # Navigation frame at top
        nav_frame = ttk.Frame(self.top)
        nav_frame.pack(fill=tk.X, padx=10, pady=5)

        self.prev_btn = ttk.Button(nav_frame, text="< Prev", command=self._prev_record)
        self.prev_btn.pack(side=tk.LEFT)

        self.next_btn = ttk.Button(nav_frame, text="Next >", command=self._next_record)
        self.next_btn.pack(side=tk.LEFT, padx=5)

        ttk.Button(nav_frame, text="|<", width=3, command=self._first_record).pack(side=tk.LEFT)
        ttk.Button(nav_frame, text=">|", width=3, command=self._last_record).pack(side=tk.LEFT, padx=5)

        self.position_label = ttk.Label(nav_frame, text="")
        self.position_label.pack(side=tk.RIGHT)

        # Content frame with Text widget and scrollbars
        content_frame = ttk.Frame(self.top)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        font_size = self.app.font_size
        self.max_name_len = max(len(col) for col in self.columns) if self.columns else 10

        # Text widget with scrollbars
        y_scroll = ttk.Scrollbar(content_frame, orient=tk.VERTICAL)
        x_scroll = ttk.Scrollbar(content_frame, orient=tk.HORIZONTAL)

        self.content_text = tk.Text(content_frame, wrap=tk.NONE,
                                   font=("Courier", font_size),
                                   bg=self.text_bg, fg=self.fg,
                                   relief="sunken", borderwidth=1,
                                   cursor="arrow",
                                   yscrollcommand=y_scroll.set,
                                   xscrollcommand=x_scroll.set)

        y_scroll.config(command=self.content_text.yview)
        x_scroll.config(command=self.content_text.xview)

        # Grid layout for text + scrollbars
        self.content_text.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        content_frame.grid_rowconfigure(0, weight=1)
        content_frame.grid_columnconfigure(0, weight=1)

        # Set tab stops for consistent alignment
        tab_width = (self.max_name_len + 2) * font_size  # pixels
        self.content_text.configure(tabs=(tab_width,))

        # Close button at bottom
        btn_frame = ttk.Frame(self.top)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Button(btn_frame, text="Close", command=self.top.destroy).pack(side=tk.RIGHT)

    def _display_record(self):
        """Display the current record in Text widget."""
        if not self.rows:
            return

        row = self.rows[self.current_index]

        # Build content and update Text widget
        self.content_text.configure(state="normal")
        self.content_text.delete("1.0", tk.END)

        for i, col in enumerate(self.columns):
            value = row[i] if i < len(row) else ""
            if value is None:
                value_str = "<NULL>"
            else:
                value_str = str(value)

            # Field name with tab for alignment
            self.content_text.insert(tk.END, f"{col}:\t{value_str}\n")

        self.content_text.configure(state="disabled")

        # Update position label
        self.position_label.config(text=f"Record {self.current_index + 1} of {len(self.rows)}")

        # Update button states
        self.prev_btn.config(state=tk.NORMAL if self.current_index > 0 else tk.DISABLED)
        self.next_btn.config(state=tk.NORMAL if self.current_index < len(self.rows) - 1 else tk.DISABLED)

        # Scroll to top
        self.content_text.yview_moveto(0)
        self.content_text.xview_moveto(0)

        # Notify callback to sync results selection
        if self.on_navigate:
            self.on_navigate(self.current_index)

    def _prev_record(self):
        """Go to previous record."""
        if self.current_index > 0:
            self.current_index -= 1
            self._display_record()

    def _next_record(self):
        """Go to next record."""
        if self.current_index < len(self.rows) - 1:
            self.current_index += 1
            self._display_record()

    def _first_record(self):
        """Go to first record."""
        self.current_index = 0
        self._display_record()

    def _last_record(self):
        """Go to last record."""
        self.current_index = len(self.rows) - 1
        self._display_record()
